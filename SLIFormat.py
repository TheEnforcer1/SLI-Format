'''
08/15/2025
File created by Anh K. Nguyen
File updated by Dany Rosete 

'''

import os
from openpyxl import load_workbook
from datetime import datetime

def main():
    print('This script will add the SLI headers and footers. Make sure file is in same folder with this script')
    
    created_name = input('Enter full name of user that created file: ')
    updated_name = input('Enter full name of user that updated file: ') 
    title = input('Enter Document Name: ')
    
    # Get the current working directory
    current_directory = os.getcwd()

    # Find all .xlsx files in the folder
    xlsx_files = [file for file in os.listdir(current_directory) if file.endswith('.xlsx')]

    if not xlsx_files:
        raise ValueError("There are no .xlsx files in the folder.")

    # Define the output file name with date in yymmdd format
    current_date = datetime.now().strftime("%y%m%d")

    # Format the date manually to m/d/yyyy
    current_date_obj = datetime.now()
    formatted_date = f"{current_date_obj.month}/{current_date_obj.day}/{current_date_obj.year}"

    # Define the updated header and footer content
    header_format = '&"Calibri"&18'

    for input_file in xlsx_files:
        file_name = str(os.path.splitext(input_file))[2:-11].strip()
        output_file = f"{file_name}_{current_date}.xlsx"
    
        # Load the workbook
        workbook = load_workbook(input_file)
    
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Update the header
            sheet.oddHeader.center.text = f"{header_format}{title} - {formatted_date}"
            sheet.oddHeader.right.text = f"Print Date: &D\nUpdate Date: {formatted_date}"
            # Update the footer
            sheet.oddFooter.left.text = f"Created Date: {formatted_date}"
            sheet.oddFooter.center.text = "&P - &N"
            sheet.oddFooter.right.text = f"Created by: {created_name}\nUpdated by: {updated_name}\n&Z&F"
    
        # Save the workbook with the updated header and footer
        workbook.save(output_file)


    print('Done')


if __name__ == '__main__':
    main()
